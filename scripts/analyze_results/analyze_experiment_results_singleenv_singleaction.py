#!/usr/bin/env python3
"""
Simple metric statistics script - supports environment, mode, and action selection; ensures sample-count consistency.
"""

import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import warnings
import sys
warnings.filterwarnings('ignore')

class SimpleMetricsStats:
    def __init__(self, outputs_dir: str = "outputs", mode: str = "fixed_point", environment: str = "pybullet", action: str = "fling"):
        self.outputs_dir = Path(outputs_dir)
        self.mode = mode
        self.environment = environment
        self.action = action  # selected action
        self.metrics_columns = [
            'chamfer_l1_sim_to_real',
            'chamfer_l2_sim_to_real',
            'chamfer_l1_real_to_sim',
            'one_sided_hausdorff_sim_to_real',
            'one_sided_hausdorff_real_to_sim',
            'sim_stability_score',
            'z_mean_error'
        ]

        # Expected environments, modes, actions, and cloth types
        self.environments = ["pybullet", "garment_dynamics", "isaacsim"]
        self.modes = ["fixed_point", "robot"]
        self.actions = ["fling", "fold", "grasp"]
        self.cloths = ["blue_dress", "green_tshirt", "grey_pleat_skirt", "white_cakeskirt", "white_shirt"]

    def collect_all_results(self) -> pd.DataFrame:
        """Collect all experiment results - pick the latest timestamp version and keep only the selected action."""
        all_results = []

        # Recursively search metrics.csv files for the chosen environment/mode/action
        pattern = f"**/{self.action}/{self.environment}/{self.mode}/**/metrics.csv"
        metrics_files = list(self.outputs_dir.glob(pattern))
        print(f"Found {len(metrics_files)} result files for action '{self.action}' environment '{self.environment}' mode '{self.mode}'")

        # Dict that stores every timestamp version per experiment
        experiment_versions = {}

        for metrics_file in metrics_files:
            try:
                # Parse experiment info from the path
                path_parts = metrics_file.parts
                if len(path_parts) >= 8:
                    cloth = path_parts[-8]
                    action = path_parts[-7]
                    environment = path_parts[-6]
                    mode = path_parts[-5]
                    robot = path_parts[-4]
                    sample = path_parts[-3]
                    timestamp = path_parts[-2]

                    # Process only the chosen mode/environment/action/cloth list
                    if mode != self.mode or environment != self.environment or action != self.action:
                        continue
                    if cloth not in self.cloths:
                        continue

                    # Build experiment key (without timestamp)
                    experiment_key = (cloth, sample)

                    # Store all versions
                    if experiment_key not in experiment_versions:
                        experiment_versions[experiment_key] = []

                    experiment_versions[experiment_key].append({
                        'timestamp': timestamp,
                        'file_path': metrics_file,
                        'mode': mode,
                        'robot': robot,
                        'environment': environment,
                        'action': action
                    })

            except Exception as e:
                print(f"Failed to parse file path {metrics_file}: {e}")
                continue

        print(f"Found {len(experiment_versions)} unique experiment combinations")

        # Pick the latest timestamp version per experiment
        for experiment_key, versions in experiment_versions.items():
            cloth, sample = experiment_key

            # Sort by timestamp and pick the latest
            versions.sort(key=lambda x: x['timestamp'], reverse=True)
            latest_version = versions[0]

            if len(versions) > 1:
                print(f"{cloth}_{sample}: selecting latest version {latest_version['timestamp']} ({len(versions)} versions total)")

            try:
                # Read metrics from the latest version
                df = pd.read_csv(latest_version['file_path'])

                if len(df) < 2:
                    print(f"Skipping due to insufficient data (need at least 2 rows): {latest_version['file_path']}")
                    continue

                # Second-to-last row is the mean; last row is variance (sqrt yields std)
                mean_row = df.iloc[-2]
                std_row = df.iloc[-1]

                # Build the result row
                result_row = {
                    'cloth': cloth,
                    'action': latest_version['action'],
                    'environment': latest_version['environment'],
                    'mode': latest_version['mode'],
                    'robot': latest_version['robot'],
                    'sample': sample,
                    'timestamp': latest_version['timestamp'],
                    'experiment_id': f"{cloth}_{latest_version['action']}_{latest_version['robot']}_{sample}",
                    'data_points': len(df) - 2  # total rows minus the statistics rows
                }

                # Extract mean and std for each metric
                for metric in self.metrics_columns:
                    if metric in df.columns:
                        mean_val = mean_row[metric] if pd.notna(mean_row[metric]) else np.nan
                        var_val = std_row[metric] if pd.notna(std_row[metric]) else np.nan

                        result_row[f'{metric}_mean'] = mean_val
                        # If variance is non-negative, sqrt to get std; otherwise NaN
                        if pd.notna(var_val) and var_val >= 0:
                            result_row[f'{metric}_std'] = np.sqrt(var_val)
                        else:
                            result_row[f'{metric}_std'] = np.nan
                    else:
                        result_row[f'{metric}_mean'] = np.nan
                        result_row[f'{metric}_std'] = np.nan

                all_results.append(result_row)
                print(f"Processed: {cloth}_{sample} (timestamp: {latest_version['timestamp']})")

            except Exception as e:
                print(f"Processing failed for {latest_version['file_path']}: {e}")
                continue

        if not all_results:
            raise ValueError(f"No valid metrics files found for action '{self.action}' environment '{self.environment}' mode '{self.mode}'!")

        result_df = pd.DataFrame(all_results)
        print(f"Total collected: {len(result_df)} experiment summaries")

        # Sort by cloth and numeric sample id
        result_df = self._sort_by_sample_number(result_df)

        # Verify sample-count consistency
        self._verify_sample_consistency(result_df)

        return result_df

    def _sort_by_sample_number(self, df: pd.DataFrame) -> pd.DataFrame:
        """Sort by cloth and numeric sample id."""
        print("Sorting by sample id...")

        # Extract the numeric part for sorting
        def extract_sample_number(sample_str):
            """Extract the numeric part from a sample string, e.g., 'sample01' or 'sample_01' -> 1"""
            import re
            match = re.search(r'(\d+)', str(sample_str))
            if match:
                return int(match.group(1))
            return 0

        # Add numeric sort column
        df['sample_number'] = df['sample'].apply(extract_sample_number)

        # Sort by cloth, sample number
        df = df.sort_values(['cloth', 'sample_number']).reset_index(drop=True)

        # Drop the temporary sort column
        df = df.drop('sample_number', axis=1)

        print("Sort complete; order: cloth -> sample id")
        return df

    def _verify_sample_consistency(self, df: pd.DataFrame):
        """Check the sample count per cloth."""
        print(f"\nVerifying sample-count consistency for action '{self.action}' environment '{self.environment}' mode '{self.mode}':")

        def extract_sample_number(sample_str):
            """Extract the numeric part from a sample string for sorting"""
            import re
            match = re.search(r'(\d+)', str(sample_str))
            if match:
                return int(match.group(1))
            return 0

        for cloth in self.cloths:
            cloth_data = df[df['cloth'] == cloth]
            sample_count = len(cloth_data)
            print(f"  {cloth}: {sample_count} samples")

            if sample_count > 0:
                # Show sample ids sorted numerically
                samples = sorted(cloth_data['sample'].tolist(), key=extract_sample_number)
                print(f"    samples: {', '.join(samples)}")
            else:
                print(f"    no data")

        print(f"\nData collection complete!")

    def create_excel_report(self, output_file: str = None):
        """Generate a simple Excel report."""
        if output_file is None:
            output_file = f"cloth_metrics_simple_{self.action}_{self.environment}_{self.mode}.xlsx"

        print("Collecting experiment data...")
        stats_df = self.collect_all_results()

        print("Creating Excel file...")

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Single worksheet: metric statistics
            stats_df.to_excel(writer, sheet_name='metric_stats', index=False)

        print(f"Excel report generated: {output_file}")

        # Apply formatting
        self._apply_formatting(output_file)

        return output_file, stats_df

    def _apply_formatting(self, excel_file):
        """Apply Excel formatting."""
        wb = openpyxl.load_workbook(excel_file)
        ws = wb['metric_stats']

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        # Set column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        # Format header row
        if ws.max_row > 0:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Number formatting (6 decimal places for small values)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value and isinstance(cell.value, (int, float)):
                    if abs(cell.value) < 1:  # decimal
                        cell.number_format = '0.000000'
                    else:  # integer or larger number
                        cell.number_format = '0.00'

        wb.save(excel_file)

    def print_summary(self, stats_df: pd.DataFrame):
        """Print a brief summary."""
        print("\n" + "="*80)
        print(f"Cloth simulation metric statistics summary")
        print("="*80)

        print(f"\nConfiguration:")
        print(f"  Action type: {self.action}")
        print(f"  Simulation environment: {self.environment}")
        print(f"  Simulation mode: {self.mode}")
        print(f"  Total experiments: {len(stats_df)}")
        print(f"  Cloth types: {stats_df['cloth'].nunique()}")

        print(f"\nCloth distribution:")
        cloth_counts = stats_df['cloth'].value_counts()
        for cloth, count in cloth_counts.items():
            print(f"  {cloth}: {count} experiments")

        print(f"\nSample id distribution:")
        sample_counts = stats_df['sample'].value_counts().sort_index()
        for sample, count in sample_counts.items():
            print(f"  {sample}: {count} cloths")

        print(f"\nData point statistics:")
        if 'data_points' in stats_df.columns:
            data_points = stats_df['data_points'].dropna()
            if len(data_points) > 0:
                print(f"  Average data points: {data_points.mean():.1f}")
                print(f"  Data point range: {data_points.min():.0f} - {data_points.max():.0f}")

        print(f"\nOverall statistics across the 7 core metrics:")
        for metric in self.metrics_columns:
            mean_col = f'{metric}_mean'
            std_col = f'{metric}_std'

            if mean_col in stats_df.columns:
                mean_values = stats_df[mean_col].dropna()
                std_values = stats_df[std_col].dropna()

                if len(mean_values) > 0:
                    print(f"  {metric}:")
                    print(f"    mean of per-experiment means: {mean_values.mean():.6f}")
                    print(f"    min of per-experiment means: {mean_values.min():.6f}")
                    print(f"    max of per-experiment means: {mean_values.max():.6f}")
                    if len(std_values) > 0:
                        print(f"    mean of per-experiment stds: {std_values.mean():.6f}")

        print("\n" + "="*80)


def select_environment():
    """Prompt the user to choose a simulation environment."""
    environments = ["pybullet", "garment_dynamics", "isaacsim"]

    print("\nPlease choose a simulation environment:")
    for i, env in enumerate(environments, 1):
        print(f"{i}. {env}")

    while True:
        try:
            choice = input(f"\nEnter your choice (1-{len(environments)}): ").strip()
            choice_idx = int(choice) - 1

            if 0 <= choice_idx < len(environments):
                return environments[choice_idx]
            else:
                print(f"Invalid choice; please enter a number between 1 and {len(environments)}")
        except ValueError:
            print(f"Please enter a valid number (1-{len(environments)})")


def select_mode():
    """Prompt the user to choose a simulation mode."""
    modes = ["fixed_point", "robot"]

    print("\nPlease choose a simulation mode:")
    for i, mode in enumerate(modes, 1):
        print(f"{i}. {mode}")

    while True:
        try:
            choice = input(f"\nEnter your choice (1-{len(modes)}): ").strip()
            choice_idx = int(choice) - 1

            if 0 <= choice_idx < len(modes):
                return modes[choice_idx]
            else:
                print(f"Invalid choice; please enter a number between 1 and {len(modes)}")
        except ValueError:
            print(f"Please enter a valid number (1-{len(modes)})")


def select_action():
    """Prompt the user to choose an action type."""
    actions = ["fling", "fold", "grasp"]

    print("\nPlease choose an action type:")
    for i, action in enumerate(actions, 1):
        print(f"{i}. {action}")

    while True:
        try:
            choice = input(f"\nEnter your choice (1-{len(actions)}): ").strip()
            choice_idx = int(choice) - 1

            if 0 <= choice_idx < len(actions):
                return actions[choice_idx]
            else:
                print(f"Invalid choice; please enter a number between 1 and {len(actions)}")
        except ValueError:
            print(f"Please enter a valid number (1-{len(actions)})")


def main():
    """Main entrypoint."""
    print("Starting cloth simulation metric analysis")

    # Prompt for environment, mode, action
    selected_environment = select_environment()
    print(f"Selected environment: {selected_environment}")

    selected_mode = select_mode()
    print(f"Selected mode: {selected_mode}")

    selected_action = select_action()
    print(f"Selected action: {selected_action}")

    # Create analyzer
    analyzer = SimpleMetricsStats(
        environment=selected_environment,
        mode=selected_mode,
        action=selected_action
    )

    try:
        # Generate report
        excel_file, stats_df = analyzer.create_excel_report()

        # Print summary
        analyzer.print_summary(stats_df)

        print(f"\nAnalysis complete!")
        print(f"Excel file: {excel_file}")
        print(f"File contents:")
        print(f"   - Worksheet 'metric_stats': mean and std of 7 metrics for the selected action")
        print(f"   - Action: {selected_action}")
        print(f"   - Environment: {selected_environment}")
        print(f"   - Mode: {selected_mode}")
        print(f"   - Sort order: cloth type, numeric sample id")
        print(f"   - Columns: cloth, action, environment, mode, robot, sample, timestamp, experiment_id, data_points")
        print(f"   - Metric columns: {', '.join([f'{m}_mean, {m}_std' for m in analyzer.metrics_columns])}")

        # Create README file
        readme_file = f"SIMPLE_METRICS_{selected_action.upper()}_{selected_environment.upper()}_{selected_mode.upper()}_README.md"
        with open(readme_file, "w", encoding="utf-8") as f:
            f.write(f"""# Cloth simulation metric statistics

## File info
- **Excel file**: {excel_file}
- **Generated at**: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}
- **Action type**: {selected_action}
- **Simulation environment**: {selected_environment}
- **Simulation mode**: {selected_mode}

## Data structure
The Excel file contains one worksheet ("metric_stats") with the following columns:

### Basic info columns
- `cloth`: cloth type (5: blue_dress, green_tshirt, grey_pleat_skirt, white_cakeskirt, white_shirt)
- `action`: action type ({selected_action})
- `environment`: simulation environment ({selected_environment})
- `mode`: simulation mode ({selected_mode})
- `robot`: robot type
- `sample`: sample id (sorted numerically)
- `timestamp`: generation timestamp (latest only)
- `experiment_id`: experiment id (cloth_action_robot_sample)
- `data_points`: valid data point count for this experiment (excluding statistics rows)

### Metric statistics columns (2 per metric)
For each metric, values are read directly from the metrics.csv file:
- `{{metric}}_mean`: mean (read from the second-to-last row of CSV)
- `{{metric}}_std`: std (sqrt of variance read from the last row of CSV)

## Data filtering and sorting
- **Action filter**: only data for the selected action ({selected_action})
- **Environment filter**: only data for the selected environment ({selected_environment})
- **Mode filter**: only data for the selected mode ({selected_mode})
- **Timestamp selection**: for multiple timestamp versions of the same experiment, pick the latest
- **Sort rule**: sorted by cloth type and numeric sample id (sample01 -> sample02 -> ...)
- **Path pattern**: `outputs/**/{selected_action}/{selected_environment}/{selected_mode}/**/metrics.csv`

## 7 core metrics (smaller is better)
1. `chamfer_l1_sim_to_real`: Chamfer L1 distance (sim -> real)
2. `chamfer_l2_sim_to_real`: Chamfer L2 distance (sim -> real)
3. `chamfer_l1_real_to_sim`: Chamfer L1 distance (real -> sim)
4. `one_sided_hausdorff_sim_to_real`: one-sided Hausdorff distance (sim -> real)
5. `one_sided_hausdorff_real_to_sim`: one-sided Hausdorff distance (real -> sim)
6. `sim_stability_score`: simulation stability score
7. `z_mean_error`: Z-axis mean error

## Data sources
- **Path pattern**: `outputs/**/{selected_action}/{selected_environment}/{selected_mode}/**/metrics.csv`
- **Mean**: read directly from the second-to-last row of each experiment's metrics.csv
- **Std**: read variance from the last row of each experiment's metrics.csv, then square root

## Usage notes
- All metrics follow 'smaller is better'
- The mean value represents the average behavior across the experiment
- The std value represents stability across the experiment (smaller = more stable)
- Each experiment uses the latest timestamp result for data consistency
- The table is sorted by cloth and numeric sample id for easy comparison
- Total rows = total experiments for this action/environment/mode (deduped by latest timestamp)
- Total columns = 9 basic info columns + 7*2=14 metric columns = 23 columns

## Selected configuration summary
- **Selected action**: {selected_action}
- **Selected environment**: {selected_environment}
- **Selected mode**: {selected_mode}
- **File naming**: cloth_metrics_simple_{{action}}_{{environment}}_{{mode}}.xlsx
- **Data path**: outputs/**/{{action}}/{{environment}}/{{mode}}/**/metrics.csv
- **Sort order**: cloth -> numeric sample id
""")

        print(f"README saved: {readme_file}")

    except Exception as e:
        print(f"Analysis failed: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()

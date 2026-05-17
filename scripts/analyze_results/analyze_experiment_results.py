#!/usr/bin/env python3
"""
Simple metric statistics script - reads precomputed mean and std values from CSV files.
"""

import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import warnings
warnings.filterwarnings('ignore')

class SimpleMetricsStats:
    def __init__(self, outputs_dir: str = "outputs"):
        self.outputs_dir = Path(outputs_dir)
        self.metrics_columns = [
            'chamfer_l1_sim_to_real',
            'chamfer_l2_sim_to_real',
            'chamfer_l1_real_to_sim',
            'one_sided_hausdorff_sim_to_real',
            'one_sided_hausdorff_real_to_sim',
            'sim_stability_score',
            'z_mean_error'
        ]

    def collect_all_results(self) -> pd.DataFrame:
        """Collect all experiment results - read precomputed statistics directly."""
        all_results = []

        # Recursively search every metrics.csv file
        metrics_files = list(self.outputs_dir.glob("**/metrics.csv"))
        print(f"Found {len(metrics_files)} result files")

        for metrics_file in metrics_files:
            try:
                # Parse experiment info from the path
                path_parts = metrics_file.parts
                if len(path_parts) >= 8:
                    cloth = path_parts[-8]
                    action = path_parts[-7]
                    environment = path_parts[-6]
                    mode = path_parts[-5]
                    robot = path_parts[-4]
                    sample = path_parts[-3]
                    timestamp = path_parts[-2]

                    # Read metrics data
                    df = pd.read_csv(metrics_file)

                    if len(df) < 2:
                        print(f"Skipping due to insufficient data (need at least 2 rows): {metrics_file}")
                        continue

                    # Second-to-last row is the mean; last row is variance (sqrt yields std)
                    mean_row = df.iloc[-2]
                    std_row = df.iloc[-1]

                    # Build the result row
                    result_row = {
                        'cloth': cloth,
                        'action': action,
                        'environment': environment,
                        'mode': mode,
                        'robot': robot,
                        'sample': sample,
                        'timestamp': timestamp,
                        'experiment_id': f"{cloth}_{action}_{robot}_{sample}",
                        'data_points': len(df) - 2  # total rows minus the statistics rows
                    }

                    # Extract mean and std for each metric
                    for metric in self.metrics_columns:
                        if metric in df.columns:
                            mean_val = mean_row[metric] if pd.notna(mean_row[metric]) else np.nan
                            var_val = std_row[metric] if pd.notna(std_row[metric]) else np.nan

                            result_row[f'{metric}_mean'] = mean_val
                            # If variance is non-negative, sqrt to get std; otherwise NaN
                            if pd.notna(var_val) and var_val >= 0:
                                result_row[f'{metric}_std'] = np.sqrt(var_val)
                            else:
                                result_row[f'{metric}_std'] = np.nan
                        else:
                            result_row[f'{metric}_mean'] = np.nan
                            result_row[f'{metric}_std'] = np.nan

                    all_results.append(result_row)
                    print(f"Processed: {cloth}_{action}_{sample} (data points: {result_row['data_points']})")

            except Exception as e:
                print(f"Processing failed for {metrics_file}: {e}")
                continue

        if not all_results:
            raise ValueError("No valid metrics files found!")

        result_df = pd.DataFrame(all_results)
        print(f"Total collected: {len(result_df)} experiment summaries")
        return result_df

    def create_excel_report(self, output_file: str = "cloth_metrics_simple.xlsx"):
        """Generate a simple Excel report."""
        print("Collecting experiment data...")
        stats_df = self.collect_all_results()

        print("Creating Excel file...")

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Single worksheet: metric statistics
            stats_df.to_excel(writer, sheet_name='metric_stats', index=False)

        print(f"Excel report generated: {output_file}")

        # Apply formatting
        self._apply_formatting(output_file)

        return output_file, stats_df

    def _apply_formatting(self, excel_file):
        """Apply Excel formatting."""
        wb = openpyxl.load_workbook(excel_file)
        ws = wb['metric_stats']

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        # Set column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        # Format header row
        if ws.max_row > 0:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Number formatting (6 decimal places for small values)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value and isinstance(cell.value, (int, float)):
                    if abs(cell.value) < 1:  # decimal
                        cell.number_format = '0.000000'
                    else:  # integer or larger number
                        cell.number_format = '0.00'

        wb.save(excel_file)

    def print_summary(self, stats_df: pd.DataFrame):
        """Print a brief summary."""
        print("\n" + "="*60)
        print("Cloth simulation metric statistics summary")
        print("="*60)

        print(f"\nBasic info:")
        print(f"  Total experiments: {len(stats_df)}")
        print(f"  Cloth types: {stats_df['cloth'].nunique()}")
        print(f"  Action types: {stats_df['action'].nunique()}")

        print(f"\nCloth distribution:")
        cloth_counts = stats_df['cloth'].value_counts()
        for cloth, count in cloth_counts.items():
            print(f"  {cloth}: {count} experiments")

        print(f"\nAction distribution:")
        action_counts = stats_df['action'].value_counts()
        for action, count in action_counts.items():
            print(f"  {action}: {count} experiments")

        print(f"\nData point statistics:")
        if 'data_points' in stats_df.columns:
            data_points = stats_df['data_points'].dropna()
            if len(data_points) > 0:
                print(f"  Average data points: {data_points.mean():.1f}")
                print(f"  Data point range: {data_points.min():.0f} - {data_points.max():.0f}")

        print(f"\nOverall statistics across the 7 core metrics:")
        for metric in self.metrics_columns:
            mean_col = f'{metric}_mean'
            std_col = f'{metric}_std'

            if mean_col in stats_df.columns:
                mean_values = stats_df[mean_col].dropna()
                std_values = stats_df[std_col].dropna()

                if len(mean_values) > 0:
                    print(f"  {metric}:")
                    print(f"    mean of per-experiment means: {mean_values.mean():.6f}")
                    print(f"    min of per-experiment means: {mean_values.min():.6f}")
                    print(f"    max of per-experiment means: {mean_values.max():.6f}")
                    if len(std_values) > 0:
                        print(f"    mean of per-experiment stds: {std_values.mean():.6f}")

        print("\n" + "="*60)


def main():
    """Main entrypoint."""
    analyzer = SimpleMetricsStats()

    try:
        # Generate report
        excel_file, stats_df = analyzer.create_excel_report()

        # Print summary
        analyzer.print_summary(stats_df)

        print(f"\nAnalysis complete!")
        print(f"Excel file: {excel_file}")
        print(f"File contents:")
        print(f"   - Worksheet 'metric_stats': mean and std of 7 metrics per experiment")
        print(f"   - Columns: cloth, action, robot, sample, experiment_id, data_points")
        print(f"   - Metric columns: {', '.join([f'{m}_mean, {m}_std' for m in analyzer.metrics_columns])}")

        # Create README file
        with open("SIMPLE_METRICS_README.md", "w", encoding="utf-8") as f:
            f.write(f"""# Cloth simulation metric statistics

## File
- **Excel file**: {excel_file}
- **Generated at**: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}

## Data structure
The Excel file contains one worksheet ("metric_stats") with the following columns:

### Basic info columns
- `cloth`: cloth type
- `action`: action type
- `robot`: robot type
- `sample`: sample id
- `experiment_id`: experiment id (cloth_action_robot_sample)
- `data_points`: valid data point count for this experiment (excluding statistics rows)

### Metric statistics columns (2 per metric)
Values are read directly from the metrics.csv file:


## 7 core metrics
1. `chamfer_l1_sim_to_real`: Chamfer L1 distance (sim -> real)
2. `chamfer_l2_sim_to_real`: Chamfer L2 distance (sim -> real)
3. `chamfer_l1_real_to_sim`: Chamfer L1 distance (real -> sim)
4. `one_sided_hausdorff_sim_to_real`: one-sided Hausdorff distance (sim -> real)
5. `one_sided_hausdorff_real_to_sim`: one-sided Hausdorff distance (real -> sim)
6. `sim_stability_score`: simulation stability score
7. `z_mean_error`: Z-axis mean error

## Data sources
- **Mean**: read directly from the second-to-last row of each experiment's metrics.csv
- **Std**: read variance from the last row of each experiment's metrics.csv, then square root

## Usage notes
- All metrics follow 'smaller is better'
- Mean value represents the average behavior across the experiment
- Std value represents stability across the experiment (smaller = more stable)
- Total rows = total experiments
- Total columns = 9 basic info columns + 7*2=14 metric columns = 23 columns
""")

        print("README saved: SIMPLE_METRICS_README.md")

    except Exception as e:
        print(f"Analysis failed: {e}")
        import traceback
        traceback.print_exc()



if __name__ == "__main__":
    main()

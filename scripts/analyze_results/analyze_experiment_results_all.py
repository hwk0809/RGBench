#!/usr/bin/env python3
"""
Environment comparison analysis script - vertical comparison of three simulation environments, with mode selection.
5 cloths x 3 actions, with 3 environments stacked vertically for each combination.
Ensures each environment has the same sample count and uses the latest result.
"""
from datetime import datetime
import pandas as pd
import numpy as np
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import warnings
warnings.filterwarnings('ignore')

def select_mode():
    """Prompt the user to choose a simulation mode."""
    print("\nPlease choose a simulation mode:")
    print("1. fixed_point")
    print("2. robot")

    while True:
        choice = input("\nEnter your choice (1-2): ").strip()
        if choice == "1":
            return "fixed_point"
        elif choice == "2":
            return "robot"
        else:
            print("Invalid choice. Please enter 1 or 2")

class VerticalEnvironmentComparisonAnalyzer:
    def __init__(self, outputs_dir: str = "outputs", mode: str = "fixed_point"):
        self.outputs_dir = Path(outputs_dir)
        self.mode = mode
        self.metrics_columns = [
            'chamfer_l1_sim_to_real',
            'chamfer_l2_sim_to_real',
            'chamfer_l1_real_to_sim',
            'one_sided_hausdorff_sim_to_real',
            'one_sided_hausdorff_real_to_sim',
            'sim_stability_score',
            'z_mean_error'
        ]

        # Expected environments, actions, and cloth types (excluding grey_sunwear and khaki_blazer)
        self.environments = ["pybullet", "garment_dynamics", "isaacsim"]
        self.actions = ["fling", "fold", "grasp"]
        self.cloths = ["blue_dress", "green_tshirt", "grey_pleat_skirt", "white_cakeskirt", "white_shirt","brown_coat", "beige_hoodie"]

    def collect_all_results(self) -> pd.DataFrame:
        """Collect all experiment results - read precomputed means and pick the latest generation time."""
        all_results = []

        # Recursively search every metrics.csv file
        metrics_files = list(self.outputs_dir.glob("**/metrics.csv"))
        print(metrics_files)
        print(f"Found {len(metrics_files)} result files")

        # Dict that stores every timestamp version per experiment
        experiment_versions = {}

        for metrics_file in metrics_files:
            try:
                # Parse experiment info from the path
                path_parts = metrics_file.parts
                if len(path_parts) >= 8:
                    cloth = path_parts[-8]
                    action = path_parts[-7]
                    environment = path_parts[-6]
                    mode = path_parts[-5]
                    robot = path_parts[-4]
                    sample = path_parts[-3]
                    timestamp = path_parts[-2]

                    # Process only the configured cloths and mode
                    if cloth not in self.cloths:
                        continue
                    if mode != self.mode:
                        continue

                    # Build experiment key
                    experiment_key = (cloth, action, environment, sample)

                    # Store every version
                    if experiment_key not in experiment_versions:
                        experiment_versions[experiment_key] = []

                    experiment_versions[experiment_key].append({
                        'timestamp': timestamp,
                        'file_path': metrics_file,
                        'mode': mode,
                        'robot': robot
                    })

            except Exception as e:
                print(f"Failed to parse file path {metrics_file}: {e}")
                continue

        print(f"Found {len(experiment_versions)} unique experiment combinations (mode: {self.mode})")

        # Pick the latest timestamp version per experiment
        for experiment_key, versions in experiment_versions.items():
            cloth, action, environment, sample = experiment_key

            # Sort by timestamp and pick the latest
            versions.sort(key=lambda x: x['timestamp'], reverse=True)
            latest_version = versions[0]

            if len(versions) > 1:
                print(f"{cloth}_{action}_{environment}_{sample}: selecting latest version {latest_version['timestamp']} ({len(versions)} versions total)")

            try:
                # Read metrics from the latest version
                df = pd.read_csv(latest_version['file_path'])

                if len(df) < 2:
                    print(f"Skipping due to insufficient data (need at least 2 rows): {latest_version['file_path']}")
                    continue

                # Second-to-last row is the mean
                mean_row = df.iloc[-2]

                # Build the result row
                result_row = {
                    'cloth': cloth,
                    'action': action,
                    'environment': environment,
                    'mode': latest_version['mode'],
                    'robot': latest_version['robot'],
                    'sample': sample,
                    'timestamp': latest_version['timestamp'],
                    'experiment_id': f"{cloth}_{action}_{latest_version['robot']}_{sample}",
                    'data_points': len(df) - 2  # total rows minus the statistics rows
                }

                # Extract mean for each metric
                for metric in self.metrics_columns:
                    if metric in df.columns:
                        mean_val = mean_row[metric] if pd.notna(mean_row[metric]) else np.nan
                        result_row[metric] = mean_val
                    else:
                        result_row[metric] = np.nan

                all_results.append(result_row)
                print(f"Processed: {environment}_{cloth}_{action}_{sample} (timestamp: {latest_version['timestamp']})")

            except Exception as e:
                print(f"Processing failed for {latest_version['file_path']}: {e}")
                continue

        if not all_results:
            raise ValueError(f"No valid metrics files found for mode '{self.mode}'!")

        result_df = pd.DataFrame(all_results)
        print(f"Total collected: {len(result_df)} experiment summaries (mode: {self.mode})")

        # Check sample-count consistency
        self._verify_sample_consistency(result_df)

        return result_df

    def _verify_sample_consistency(self, df: pd.DataFrame):
        """Check that the sample counts per environment are consistent."""
        print(f"\nVerifying sample-count consistency for mode '{self.mode}':")

        for cloth in self.cloths:
            for action in self.actions:
                print(f"\n  {cloth}_{action}:")
                sample_counts = {}

                for env in self.environments:
                    env_data = df[
                        (df['cloth'] == cloth) &
                        (df['action'] == action) &
                        (df['environment'] == env)
                    ]
                    sample_counts[env] = len(env_data)
                    print(f"    {env}: {len(env_data)} samples")

                # Check for consistency
                counts = list(sample_counts.values())
                if len(set(counts)) > 1:
                    print(f"    inconsistent sample counts: {sample_counts}")
                else:
                    print(f"    consistent sample counts: {counts[0]}")

    def create_vertical_comparison_table(self, df: pd.DataFrame) -> pd.DataFrame:
        """Create the vertical environment comparison table."""
        comparison_rows = []

        # For each cloth-action combination, create one row per environment
        for cloth in self.cloths:
            for action in self.actions:
                for env in self.environments:
                    # Filter to the current cloth-action-environment
                    env_data = df[
                        (df['cloth'] == cloth) &
                        (df['action'] == action) &
                        (df['environment'] == env)
                    ]

                    # Build the row
                    row = {
                        'cloth': cloth,
                        'action': action,
                        'environment': env,
                        'mode': self.mode,
                        'combination': f"{cloth}_{action}",
                        'samples': len(env_data)
                    }

                    if len(env_data) > 0:
                        # Average across all samples
                        for metric in self.metrics_columns:
                            if metric in env_data.columns:
                                # Mean across samples
                                metric_values = env_data[metric].dropna()
                                if len(metric_values) > 0:
                                    avg_value = metric_values.mean()
                                    row[metric] = avg_value
                                else:
                                    row[metric] = np.nan
                            else:
                                row[metric] = np.nan
                    else:
                        # No data
                        for metric in self.metrics_columns:
                            row[metric] = np.nan

                    comparison_rows.append(row)

        return pd.DataFrame(comparison_rows)

    def create_excel_report(self, output_file: str = None):
        """Generate the vertical environment comparison Excel report."""
        print(output_file)
        # Auto-generate a filename only when one is not provided
        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"environment_comparison_vertical_{timestamp}.xlsx"
        # If main passes an absolute path, use it as-is
        print("Collecting experiment data...")
        raw_df = self.collect_all_results()
        print("Creating vertical comparison table...")
        comparison_df = self.create_vertical_comparison_table(raw_df)
        comparison_df = comparison_df.sort_values(['cloth', 'action', 'environment'])
        print("Creating Excel file...")
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            comparison_df.to_excel(writer, sheet_name='env_vertical_compare', index=False)
            summary_df = self.create_summary_statistics(comparison_df)
            summary_df.to_excel(writer, sheet_name='env_summary', index=False)
            best_env_df = self.create_best_environment_table(comparison_df)
            best_env_df.to_excel(writer, sheet_name='best_env_per_metric', index=False)
            consistency_df = self.create_sample_consistency_table(raw_df)
            consistency_df.to_excel(writer, sheet_name='sample_consistency', index=False)
        print(f"Excel report generated: {output_file}")
        self._apply_formatting(output_file)
        return output_file, comparison_df, summary_df, best_env_df, consistency_df

    def create_sample_consistency_table(self, df: pd.DataFrame) -> pd.DataFrame:
        """Create the sample-consistency table."""
        consistency_rows = []

        for cloth in self.cloths:
            for action in self.actions:
                row = {
                    'cloth': cloth,
                    'action': action,
                    'mode': self.mode,
                    'combination': f"{cloth}_{action}"
                }

                # Count samples per environment
                for env in self.environments:
                    env_data = df[
                        (df['cloth'] == cloth) &
                        (df['action'] == action) &
                        (df['environment'] == env)
                    ]
                    row[f'{env}_samples'] = len(env_data)

                    # Record timestamp range
                    if len(env_data) > 0:
                        timestamps = env_data['timestamp'].tolist()
                        row[f'{env}_timestamps'] = ', '.join(sorted(set(timestamps)))
                    else:
                        row[f'{env}_timestamps'] = 'None'

                # Check consistency
                sample_counts = [row[f'{env}_samples'] for env in self.environments]
                row['is_consistent'] = len(set(sample_counts)) == 1
                row['min_samples'] = min(sample_counts)
                row['max_samples'] = max(sample_counts)

                consistency_rows.append(row)

        return pd.DataFrame(consistency_rows)

    def create_summary_statistics(self, comparison_df: pd.DataFrame) -> pd.DataFrame:
        """Create the summary statistics table."""
        summary_rows = []

        for env in self.environments:
            env_data = comparison_df[comparison_df['environment'] == env]

            for metric in self.metrics_columns:
                if metric in env_data.columns:
                    values = env_data[metric].dropna()
                    if len(values) > 0:
                        summary_row = {
                            'environment': env,
                            'mode': self.mode,
                            'metric': metric,
                            'mean': values.mean(),
                            'std': values.std(),
                            'min': values.min(),
                            'max': values.max(),
                            'count': len(values)
                        }
                        summary_rows.append(summary_row)

        return pd.DataFrame(summary_rows)
    def create_best_environment_table(self, comparison_df: pd.DataFrame) -> pd.DataFrame:
        """Create the best-environment-per-metric table."""
        best_env_rows = []

        for metric in self.metrics_columns:
            # Find the best environment for each cloth-action combination
            for cloth in self.cloths:
                for action in self.actions:
                    combo_data = comparison_df[
                        (comparison_df['cloth'] == cloth) &
                        (comparison_df['action'] == action)
                    ]

                    if len(combo_data) > 0 and metric in combo_data.columns:
                        # Pick the environment with the minimum metric value (smaller is better)
                        valid_data = combo_data.dropna(subset=[metric])
                        if len(valid_data) > 0:
                            best_idx = valid_data[metric].idxmin()
                            best_row = valid_data.loc[best_idx]

                            best_env_rows.append({
                                'cloth': cloth,
                                'action': action,
                                'mode': self.mode,
                                'metric': metric,
                                'best_environment': best_row['environment'],
                                'best_value': best_row[metric],
                                'combination': f"{cloth}_{action}"
                            })

        return pd.DataFrame(best_env_rows)

    def _apply_formatting(self, excel_file):
        """Apply Excel formatting."""
        wb = openpyxl.load_workbook(excel_file)

        # Format main comparison sheet
        if 'env_vertical_compare' in wb.sheetnames:
            ws = wb['env_vertical_compare']
            self._format_main_comparison_sheet(ws)

        # Format summary sheet
        if 'env_summary' in wb.sheetnames:
            ws = wb['env_summary']
            self._format_summary_sheet(ws)

        # Format best-environment sheet
        if 'best_env_per_metric' in wb.sheetnames:
            ws = wb['best_env_per_metric']
            self._format_best_env_sheet(ws)

        # Format sample-consistency sheet
        if 'sample_consistency' in wb.sheetnames:
            ws = wb['sample_consistency']
            self._format_consistency_sheet(ws)

        wb.save(excel_file)

    def _format_consistency_sheet(self, ws):
        """Format the sample-consistency worksheet."""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        consistent_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # green
        inconsistent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # red

        # Set column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 30)

        # Format header row
        if ws.max_row > 0:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Format data rows
        for row in ws.iter_rows(min_row=2):
            # Locate the is_consistent cell
            is_consistent_cell = None
            for cell in row:
                if ws[f"{cell.column_letter}1"].value == 'is_consistent':
                    is_consistent_cell = cell
                    break

            # Color by consistency
            if is_consistent_cell and is_consistent_cell.value is not None:
                fill_color = consistent_fill if is_consistent_cell.value else inconsistent_fill
                for cell in row:
                    cell.fill = fill_color

                    # Number formatting
                    if cell.value and isinstance(cell.value, (int, float)):
                        cell.number_format = '0'

    def _format_main_comparison_sheet(self, ws):
        """Format the main comparison worksheet."""
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        # Environment colors
        env_colors = {
            'pybullet': PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid"),
            'garment_dynamics': PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid"),
            'isaacsim': PatternFill(start_color="FFF2E8", end_color="FFF2E8", fill_type="solid")
        }

        # Set column widths
        ws.column_dimensions['A'].width = 18  # cloth
        ws.column_dimensions['B'].width = 12  # action
        ws.column_dimensions['C'].width = 16  # environment
        ws.column_dimensions['D'].width = 16  # mode
        ws.column_dimensions['E'].width = 20  # combination
        ws.column_dimensions['F'].width = 10  # samples

        # Metric column widths
        for i, metric in enumerate(self.metrics_columns):
            col_letter = chr(ord('G') + i)  # G, H, I, ...
            ws.column_dimensions[col_letter].width = 20

        # Format header row
        if ws.max_row > 0:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Collect rows so we can highlight the per-combination minimum
        data_rows = list(ws.iter_rows(min_row=2))

        # Process rows grouped by combination
        current_combo = None
        combo_rows = []

        for row_idx, row in enumerate(data_rows):
            combination = row[4].value  # combination column (5th)

            if current_combo != combination:
                # Highlight values in the previous combination
                if combo_rows:
                    self._highlight_best_values_in_combo(combo_rows, env_colors)

                # Start a new combination
                current_combo = combination
                combo_rows = [row]
            else:
                combo_rows.append(row)

        # Handle the last combination
        if combo_rows:
            self._highlight_best_values_in_combo(combo_rows, env_colors)

        # Format data rows
        prev_combination = None
        for row_num, row in enumerate(data_rows, start=2):
            # Get current combination info
            current_combination = row[4].value  # combination column

            # Add a thick border between different combinations
            if prev_combination is not None and current_combination != prev_combination:
                for cell in row:
                    cell.border = Border(top=Side(style='thick'))

            # Update previous combination
            prev_combination = current_combination

    def _highlight_best_values_in_combo(self, combo_rows, env_colors):
        """Within one combination's 3 rows, bold the minimum value for each metric."""
        # Metric columns start at the 7th column (G) because mode was added
        metric_start_col = 6  # 0-based index; column G is index 6

        for metric_idx, metric in enumerate(self.metrics_columns):
            col_idx = metric_start_col + metric_idx

            # Collect this metric's value across the three environments
            values = []
            valid_rows = []

            for row in combo_rows:
                if col_idx < len(row):
                    cell = row[col_idx]
                    if cell.value and isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                        values.append(cell.value)
                        valid_rows.append(row)

            # Find the minimum value
            if values:
                min_value = min(values)

                # Style the cell that holds the minimum
                for row in valid_rows:
                    if col_idx < len(row):
                        cell = row[col_idx]
                        if cell.value == min_value:
                            # Read environment info for coloring
                            env_cell = row[2]  # environment column
                            env_name = env_cell.value

                            # Number formatting
                            if abs(cell.value) < 1:
                                cell.number_format = '0.000000'
                            else:
                                cell.number_format = '0.00'

                            # Background color by environment
                            if env_name in env_colors:
                                cell.fill = env_colors[env_name]

                            # Bold the minimum
                            cell.font = Font(bold=True, size=11)
                            break  # only bold the first minimum (in case of ties)

                # Style the non-minimum cells with normal formatting
                for row in valid_rows:
                    if col_idx < len(row):
                        cell = row[col_idx]
                        if cell.value != min_value:
                            # Read environment info for coloring
                            env_cell = row[2]  # environment column
                            env_name = env_cell.value

                            # Number formatting
                            if abs(cell.value) < 1:
                                cell.number_format = '0.000000'
                            else:
                                cell.number_format = '0.00'

                            # Background color by environment
                            if env_name in env_colors:
                                cell.fill = env_colors[env_name]

                            # Normal font
                            cell.font = Font(bold=False, size=11)

        # Style the non-metric columns with normal formatting
        for row in combo_rows:
            # Read environment info for coloring
            env_cell = row[2]  # environment column
            env_name = env_cell.value

            # Handle the first 6 columns (cloth, action, environment, mode, combination, samples)
            for col_idx in range(min(6, len(row))):
                cell = row[col_idx]

                # Background color by environment
                if env_name in env_colors:
                    cell.fill = env_colors[env_name]

                # Normal font
                cell.font = Font(bold=False, size=11)

                # Number formatting (if numeric)
                if cell.value and isinstance(cell.value, (int, float)):
                    if abs(cell.value) < 1:
                        cell.number_format = '0.000000'
                    else:
                        cell.number_format = '0.00'

    def _format_summary_sheet(self, ws):
        """Format the summary worksheet."""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        # Set column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 25)

        # Format header row
        if ws.max_row > 0:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Number formatting
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value and isinstance(cell.value, (int, float)):
                    if abs(cell.value) < 1:
                        cell.number_format = '0.000000'
                    else:
                        cell.number_format = '0.00'

    def _format_best_env_sheet(self, ws):
        """Format the best-environment worksheet."""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        # Set column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 22)

        # Format header row
        if ws.max_row > 0:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Number formatting and environment colors
        env_colors = {
            'pybullet': PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid"),
            'garment_dynamics': PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid"),
            'isaacsim': PatternFill(start_color="FFF2E8", end_color="FFF2E8", fill_type="solid")
        }

        for row in ws.iter_rows(min_row=2):
            # Read best-environment info
            best_env = row[4].value  # best_environment column (adjusted index)

            for cell in row:
                if cell.value and isinstance(cell.value, (int, float)):
                    if abs(cell.value) < 1:
                        cell.number_format = '0.000000'
                    else:
                        cell.number_format = '0.00'

                # Background color by best environment
                if best_env in env_colors:
                    cell.fill = env_colors[best_env]

    def print_summary(self, comparison_df: pd.DataFrame, summary_df: pd.DataFrame, best_env_df: pd.DataFrame, consistency_df: pd.DataFrame):
        """Print the analysis summary."""
        print("\n" + "="*80)
        print(f"Simulation environment vertical comparison summary (mode: {self.mode})")
        print("="*80)

        print(f"\nBasic info:")
        print(f"  Simulation mode: {self.mode}")
        print(f"  Cloth types: {len(self.cloths)} ({', '.join(self.cloths)})")
        print(f"  Action types: {len(self.actions)} ({', '.join(self.actions)})")
        print(f"  Simulation environments: {len(self.environments)} ({', '.join(self.environments)})")
        print(f"  Total comparison rows: {len(comparison_df)} (15 combinations x 3 environments)")

        print(f"\nSample consistency check:")
        consistent_combos = consistency_df[consistency_df['is_consistent'] == True]
        inconsistent_combos = consistency_df[consistency_df['is_consistent'] == False]

        print(f"  Combinations with consistent sample counts: {len(consistent_combos)}/{len(consistency_df)}")
        print(f"  Combinations with inconsistent sample counts: {len(inconsistent_combos)}/{len(consistency_df)}")

        if len(inconsistent_combos) > 0:
            print(f"\n  Inconsistent combinations:")
            for _, row in inconsistent_combos.iterrows():
                print(f"    - {row['combination']}: min={row['min_samples']}, max={row['max_samples']}")

        print(f"\nData completeness check:")
        for env in self.environments:
            env_data = comparison_df[comparison_df['environment'] == env]
            total_samples = env_data['samples'].sum()
            missing_combos = (env_data['samples'] == 0).sum()
            print(f"  {env}: {total_samples} samples, {missing_combos}/{len(self.cloths)*len(self.actions)} missing combinations")

        print(f"\nOverall performance ranking per environment:")
        for metric in self.metrics_columns:
            print(f"\n  {metric}:")
            metric_summary = summary_df[summary_df['metric'] == metric].copy()
            if len(metric_summary) > 0:
                metric_summary = metric_summary.sort_values('mean')
                for i, (_, row) in enumerate(metric_summary.iterrows(), 1):
                    print(f"    {i}. {row['environment']}: {row['mean']:.6f} +/- {row['std']:.6f}")

        print(f"\nWins per environment:")
        best_env_counts = best_env_df['best_environment'].value_counts()
        for env, count in best_env_counts.items():
            percentage = count / len(best_env_df) * 100
            print(f"  {env}: {count} ({percentage:.1f}%)")

        print("\n" + "="*80)


def main():
    """Main entrypoint."""
    # Prompt for mode
    selected_mode = select_mode()

    try:
        # Locate project root
        script_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(script_dir)
        # Save directory is outputs/analysis/all
        save_dir = os.path.join(project_root, "outputs", "analysis", "all")
        os.makedirs(save_dir, exist_ok=True)

        # Data directory remains outputs
        outputs_dir = os.path.join(project_root, "outputs")
        analyzer = VerticalEnvironmentComparisonAnalyzer(outputs_dir=outputs_dir, mode=selected_mode)

        # Auto-generate a filename with timestamp
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        save_path = os.path.join(save_dir, f"environment_comparison_vertical_{timestamp}.xlsx")

        excel_file, comparison_df, summary_df, best_env_df, consistency_df = analyzer.create_excel_report(output_file=save_path)

        analyzer.print_summary(comparison_df, summary_df, best_env_df, consistency_df)

        print(f"\nVertical environment comparison analysis complete!")
        print(f"Excel file: {excel_file}")
        print(f"File contains four worksheets:")
        print(f"   1. 'env_vertical_compare': three environments stacked per cloth-action combination")
        print(f"   2. 'env_summary': per-environment overall statistics per metric")
        print(f"   3. 'best_env_per_metric': best environment per combination per metric")
        print(f"   4. 'sample_consistency': verifies sample counts across environments")
        print(f"\nSample consistency guarantee (mode: {selected_mode}):")
        print(f"   Only data for mode '{selected_mode}' is analyzed")
        print(f"   Each experiment uses the latest timestamp result")
        print(f"   The same cloth-action combination uses the same samples across environments")
        print(f"   Detailed consistency results live in the 4th worksheet")

    except Exception as e:
        print(f"Analysis failed: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
